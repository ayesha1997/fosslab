import openpyxl
import cx_Oracle

#Oracle connection starts here
connection = cx_Oracle.connect('system/shaquib@localhost/XE')
cursor = connection.cursor()
print("Database version:", connection.version)
print(cx_Oracle.version)
print(connection.current_schema)


from sys import modules   

from openpyxl import Workbook
wb = openpyxl.load_workbook('C:/Users/AbiShaquib/Desktop/CTS/SQL/Book1.xlsx',data_only=True)
ws = wb['Sheet3']

x=1
m=1

# looping through each column
for i in range(3, ws.max_row+1):       
    Book_id = ws.cell(row=i, column=1).value  
    Name = ws.cell(row=i, column=2).value
    Price = ws.cell(row=i, column=3).value

    insert_table = "INSERT INTO book (Book_id, Name, Price)" + " VALUES (:1, :2, :3)"

    cursor.execute(insert_table, (Book_id, Name, Price))
    connection.commit()

connection.close()
